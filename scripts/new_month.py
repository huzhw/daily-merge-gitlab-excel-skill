# -*- coding: utf-8 -*-
"""月初脚本：从上月 Excel 复制未完成任务 + 写入当天 md 新需求（按仓库分组合并），序号跨月延续"""
import openpyxl, os, re, copy
import sys
from datetime import datetime
from shared import find_report_dir, format_desc, to_chinese, SEP, is_temp_task, next_workday
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Windows 控制台/管道默认 GBK，强制 UTF-8 输出避免中文乱码
try:
    sys.stdout.reconfigure(encoding='utf-8')
except (AttributeError, ValueError, OSError):
    pass

DESKTOP = r"C:\Users\Administrator\Desktop"
TODAY = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
Y = TODAY.strftime("%Y"); M = TODAY.strftime("%m"); MM = TODAY.strftime("%m"); DD = TODAY.strftime("%d")

REPO_MAP = {
    'lanxum-amisp': '档案V6', 'lanxum-amisp-java': '档案V6', 'lanxum-amisp-react': '档案V6',
    'workingpaper-v5.5': '中信底稿V5',
    'standard_thdg_zxdm': '中信底稿V5',
    '中信底稿V5': '中信底稿V5',
    '中信底稿v5': '中信底稿V5',
    '档案V6': '档案V6',
    '档案系统V6': '档案V6',
    '智能数据底座': '智能数据底座',
}

RD = find_report_dir(DESKTOP, Y, M)
MD = os.path.join(RD, f"日报需求记录-{Y}-{MM}-{DD}.md")
XL = os.path.join(RD, f"日报表格-胡志伟~~{MM}-{DD}.xlsx")


def parse_date_val(val):
    """统一解析单元格日期值 → date 对象"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    try:
        return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
    except:
        return None


def parse_pct_val(val):
    """解析 E 列完成百分比 → 0~1 浮点数"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace('%', '').strip()
    try:
        return float(s) / 100.0
    except:
        return 0


def find_last_month_xlsx():
    """找上月目录下日期最新的 Excel 文件，跨年自动处理。返回 (路径, 文件名) 或 (None, None)"""
    if TODAY.month == 1:
        last_year = str(TODAY.year - 1)
        last_month = '12'
    else:
        last_year = Y
        last_month = f"{TODAY.month - 1:02d}"

    last_rd = find_report_dir(DESKTOP, last_year, last_month)
    if not os.path.exists(last_rd):
        return None, None

    pattern = re.compile(r'日报表格-胡志伟~~(\d{2})-(\d{2})\.xlsx$')
    best_path = None
    best_date = None
    for fname in os.listdir(last_rd):
        m = pattern.match(fname)
        if not m:
            continue
        try:
            f_date = datetime(int(last_year), int(m.group(1)), int(m.group(2)))
        except ValueError:
            continue
        if best_date is None or f_date > best_date:
            best_date = f_date
            best_path = os.path.join(last_rd, fname)
    return best_path, best_date


def get_last_month_data(xlsx_path):
    """读上月最后一天 Excel，返回 (max_seq, unfinished_rows, last_g_date, remaining_hours)

    unfinished_rows: list[{repo, desc, pct, hours, ai_h}]
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    # ── 找到填充说明行 ──
    notes_row = ws.max_row + 1
    for row in range(2, ws.max_row + 1):
        a = ws.cell(row=row, column=1).value
        if a and '填充说明' in str(a):
            notes_row = row
            break

    # ── 构建 B 列回溯表（每行对应的真实项目名）──
    repo_by_row = {}
    prev_repo = None
    for row in range(2, notes_row):
        bv = ws.cell(row=row, column=2).value
        if bv and str(bv).strip():
            prev_repo = str(bv).strip()
        repo_by_row[row] = prev_repo

    # ── 找最大序号 ──
    max_seq = 0
    for row in range(2, notes_row):
        cv = ws.cell(row=row, column=3).value
        if cv is not None and str(cv).strip().isdigit():
            max_seq = max(max_seq, int(str(cv).strip()))

    # ── 找最后一个有 G 列值的行，计算当天剩余工时 ──
    last_g = TODAY
    last_g_row = 0
    for row in range(2, notes_row):
        cv = ws.cell(row=row, column=3).value
        if cv is None or not str(cv).strip().isdigit():
            continue
        gv = ws.cell(row=row, column=7).value
        if gv is not None:
            try:
                if isinstance(gv, datetime):
                    last_g = gv
                else:
                    last_g = datetime.strptime(str(gv)[:10], '%Y-%m-%d')
                last_g_row = row
            except:
                pass

    day_hours = 0
    for row in range(last_g_row, 1, -1):
        rd = parse_date_val(ws.cell(row=row, column=7).value)
        if rd is None or rd != last_g.date():
            break
        # 跳过当天创建当天完成的临时任务（开会/部署/沟通）
        if is_temp_task(ws.cell(row=row, column=4).value,
                        ws.cell(row=row, column=6).value,
                        ws.cell(row=row, column=7).value,
                        ws.cell(row=row, column=10).value):
            continue
        hv = ws.cell(row=row, column=8).value
        if hv:
            try:
                day_hours += float(hv)
            except:
                pass
    remaining = day_hours % 8

    # ── 找最后一个日期块中未完成的行 ──
    date_blocks = []
    i = 2
    while i < notes_row:
        av = ws.cell(row=i, column=1).value
        d = parse_date_val(av)
        if d is not None:
            j = i
            while j + 1 < notes_row:
                next_a = ws.cell(row=j + 1, column=1).value
                if next_a is not None:
                    break
                j += 1
            date_blocks.append((d, i, j))
            i = j + 1
        else:
            i += 1

    if not date_blocks:
        return max_seq, [], last_g, remaining

    _, block_start, block_end = date_blocks[-1]

    unfinished = []
    for row in range(block_start, block_end + 1):
        cv = ws.cell(row=row, column=3).value
        if cv is None or not str(cv).strip().isdigit():
            continue
        ev = ws.cell(row=row, column=5).value
        pct = parse_pct_val(ev)
        if pct >= 1:
            continue

        repo = repo_by_row.get(row, '')
        desc = ws.cell(row=row, column=4).value
        hours = float(ws.cell(row=row, column=8).value or 0)

        # G 列保留原值
        gv = ws.cell(row=row, column=7).value
        g_date = None
        if gv:
            if isinstance(gv, datetime):
                g_date = gv
            else:
                try:
                    g_date = datetime.strptime(str(gv)[:10], '%Y-%m-%d')
                except:
                    pass

        # F 列（任务创建时间）保留原值
        fv = ws.cell(row=row, column=6).value
        f_date = None
        if fv:
            if isinstance(fv, datetime):
                f_date = fv
            else:
                try:
                    f_date = datetime.strptime(str(fv)[:10], '%Y-%m-%d')
                except:
                    pass

        # N 列（保留原值，不重新格式化）
        nv = ws.cell(row=row, column=14).value
        ai_note = str(nv).strip() if nv else ''

        unfinished.append({
            'repo': repo,
            'seq': int(str(cv).strip()),
            'desc': desc,
            'pct': ev if ev is not None else '0%',
            'hours': hours,
            'g_date': g_date,
            'f_date': f_date,
            'ai_note': ai_note,
        })

    wb.close()
    return max_seq, unfinished, last_g, remaining


def load_template(xlsx_path):
    """从基准 Excel 提取表头(第1行)和最后一个数据行的单元格样式，作为新表样式模板"""
    twb = openpyxl.load_workbook(xlsx_path)
    tws = twb[twb.sheetnames[0]]

    header = {}
    for col in range(1, 15):
        cell = tws.cell(row=1, column=col)
        header[col] = {
            'font': copy.copy(cell.font), 'fill': copy.copy(cell.fill),
            'border': copy.copy(cell.border), 'alignment': copy.copy(cell.alignment),
        }
    header['height'] = tws.row_dimensions[1].height or 57

    data = {}
    data_row = None
    # 取最后一个日期块的锚行（A 列有日期值 + 有 C 序号），样式最完整（边框/数字格式齐全）
    for r in range(tws.max_row, 1, -1):
        av = tws.cell(row=r, column=1).value
        cv = tws.cell(row=r, column=3).value
        if av is not None and str(av).strip() and cv is not None and str(cv).strip().isdigit():
            data_row = r
            break
    if data_row:
        for col in range(1, 15):
            cell = tws.cell(row=data_row, column=col)
            data[col] = {
                'font': copy.copy(cell.font), 'fill': copy.copy(cell.fill),
                'border': copy.copy(cell.border), 'alignment': copy.copy(cell.alignment),
                'number_format': cell.number_format,
            }

    twb.close()
    return header, data


def parse():
    """解析当天 md，按仓库分组合并已完成任务"""
    ts_by_repo = {}
    repo_order = []
    if not os.path.exists(MD):
        return []
    with open(MD, encoding='utf-8') as f:
        c = f.read()
    for l in c.split('\n'):
        l = l.strip()
        if not l.startswith('|') or '---' in l or '序号' in l or '空行' in l:
            continue
        p = [x.strip() for x in l.split('|')[1:-1]]
        if len(p) < 9:
            continue
        seq, date, repo, desc, modules, status, human_h, ai_h, note = p[:9]
        if status not in ('已完成', '100%') or not seq.isdigit():
            continue

        display = REPO_MAP.get(repo, repo).replace('、', '+')
        if display not in ts_by_repo:
            ts_by_repo[display] = []
            repo_order.append(display)
        ts_by_repo[display].append({
            'desc': desc,
            'modules': modules,
            'note': note,
            'human_h': float(human_h.rstrip('h') or 0),
            'ai_h': float(ai_h.rstrip('h') or 0),
        })

    result = []
    for repo in repo_order:
        tasks = ts_by_repo[repo]
        desc_parts = []
        for i, t in enumerate(tasks, 1):
            part = f"{to_chinese(i)}、{format_desc(t['desc'])}"
            if t['modules'] and t['modules'] != '—':
                part += f"\n{t['modules']}"
            if t['note'] and t['note'] != '—':
                part += f"\n{t['note']}"
            desc_parts.append(part)

        result.append({
            'repo': repo,
            'desc': ('\n' + SEP + '\n').join(desc_parts),
            'human_h': sum(t['human_h'] for t in tasks),
            'ai_h': sum(t['ai_h'] for t in tasks),
        })

    return result


# ── 共用样式构建 ──
def make_styles():
    hfont = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='1450B8', end_color='1450B8', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
    dfont = Font(name='Microsoft YaHei', size=11)
    dalign = Alignment(horizontal='center', vertical='center')
    bthin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    bdata = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    return hfont, hfill, halign, dfont, dalign, bthin, bdata


def write_header(ws, styles, header_t=None):
    """表头。header_t 为基准表样式模板时逐列套用，否则用默认样式"""
    hfont, hfill, halign, _, _, bthin, _ = styles
    hd = ['*日期', '*项目名称', '*序号\n（全局唯一）', '*任务描述', '*完成百分比',
          '*任务创建时间\n（录入后勿修改）', '*预计完成时间\n（录入后勿修改）',
          '*预计工作人时\n', '调整预计完成时间\n（由于优先级问题调整时填写）',
          '实际完成时间', '*实际工作人时', '插队序号\n若有插队必填', '延期/调整原因', '备注/说明']
    for i, h in enumerate(hd, 1):
        c = ws.cell(row=1, column=i, value=h)
        if header_t and i in header_t:
            st = header_t[i]
            c.font = copy.copy(st['font']); c.fill = copy.copy(st['fill'])
            c.border = copy.copy(st['border']); c.alignment = copy.copy(st['alignment'])
        else:
            c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = bthin
    ws.row_dimensions[1].height = header_t.get('height', 57) if header_t else 57


def write_notes(ws, data_row_count, styles):
    """填充说明行，字体统一 Microsoft YaHei 9 号"""
    nfont = Font(name='Microsoft YaHei', size=9)
    notes = [
        ('填充说明：\n填写时例子请删除', 1),
        ('1、带*号的字段必填。', 2),
        ('2、前一天已完成的任务不需要重新出现在当天的记录中（若之前的任务完成后有bug需要修改，需要新建一条记录，复制之前任务的描述后面增加【bug】字样，并在插队序号中填写之前的任务序号，主序号依然递增）', 2),
        ('3、如果遇到更优先级的任务阻断了当前任务，需要新增阻断任务的信息，在插队序号中写入被影响的任务序号，多个任务逗号隔开，并插入/修改被影响任务的【调整预计完成时间】字段', 2),
        ('4、工时采用小时单位，按一天8小时计算，其中实际工时每天更新一直到任务完成100%，任务完成的当天做最后一次记录更新后，第二天不用重复写入该记录。', 2),
        ('5、此表在一个自然月内持续迭代，下一个自然月后重新起一个新表。', 2),
        ('6、备注按需填写，延迟原因需要简要写明。', 2),
        ('7、每天日报发送到公邮：kpi-daily-report@lscjz.com；每人每天一封。', 2),
        ('8、开会和沟通也当作任务记录时长', 2),
    ]
    nr = data_row_count + 5
    for i, (txt, col) in enumerate(notes):
        c = ws.cell(row=nr + i, column=col, value=txt)
        c.font = nfont


def write_row(ws, row, seq, repo, desc, pct, hours, g_date, ai_note, styles, f_date=None, tdata=None):
    """写一行数据。ai_note 为空时不写 N 列。f_date 为空时 F 列填今天。
    样式优先套用基准表模板 tdata（含数字格式），否则用默认样式。"""
    _, _, _, dfont, dalign, _, bdata = styles

    def style_cell(col, value=None, numfmt=None):
        cell = ws.cell(row=row, column=col)
        if tdata and col in tdata:
            st = tdata[col]
            cell.font = copy.copy(st['font'])
            cell.border = copy.copy(st['border'])
        else:
            cell.font = dfont
            cell.border = bdata
        nf = numfmt
        if nf is None and tdata and col in tdata:
            nf = tdata[col].get('number_format')
        if nf and nf != 'General':
            cell.number_format = nf
        if value is not None:
            cell.value = value
        # D/M/N 左对齐，其余居中
        if col in (4, 13, 14):
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return cell

    style_cell(1, TODAY, 'yyyy/m/d;@')
    if repo:
        style_cell(2, repo)
    style_cell(3, seq)
    style_cell(4, desc)
    ws.row_dimensions[row].height = 35
    # E 列统一写数字 0（未完成/新任务均为 0%），0% 数字格式
    style_cell(5, 0, '0%')
    style_cell(6, f_date or TODAY, 'yyyy/m/d;@')
    style_cell(7, g_date, 'yyyy/m/d;@')
    style_cell(8, hours)
    if ai_note:
        style_cell(14, ai_note)
    # I/J/K/L/M 列无值，但套用样式（I/J/K/L 居中，M 左对齐）
    for c in (9, 10, 11, 12, 13):
        style_cell(c)


def finalize(ws, data_start, data_end):
    """A 列合并、列宽、冻结、保存"""
    if data_end > data_start:
        ws.merge_cells(start_row=data_start, start_column=1, end_row=data_end, end_column=1)
    for r in range(data_start, data_end + 1):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')

    cw = {'A': 13, 'B': 23, 'C': 15, 'D': 60, 'E': 16, 'F': 18, 'G': 18, 'H': 14,
          'I': 32, 'J': 13, 'K': 15, 'L': 14, 'M': 25, 'N': 60}
    for k, v in cw.items():
        ws.column_dimensions[k].width = v
    ws.freeze_panes = 'A2'
    ws.sheet_view.topLeftCell = 'A1'
    wb.save(XL)


def main():
    os.makedirs(RD, exist_ok=True)
    styles = make_styles()

    # ── 1. 读上月数据 ──
    last_xlsx, last_date = find_last_month_xlsx()
    max_seq = 0
    unfinished = []
    gd = TODAY
    gr = 0

    if last_xlsx:
        print(f"上月基准: {os.path.basename(last_xlsx)} ({last_date})")
        max_seq, unfinished, last_g, remaining = get_last_month_data(last_xlsx)
        gd = last_g
        gr = remaining
        print(f"  最大序号: {max_seq}, 未完成: {len(unfinished)} 行, G={gd.strftime('%m-%d')} 剩余{gr}h")
    else:
        print("未找到上月 Excel，序号从 1 开始")

    # ── 2. 解析当天 md ──
    new_tasks = parse()
    if new_tasks:
        print(f"md 新任务: {len(new_tasks)} 个仓库分组")
    else:
        print("md 无已完成任务")

    total = len(unfinished) + len(new_tasks)
    if total == 0:
        print("无未完成任务、无新任务，跳过")
        return

    # ── 3. 创建新 Workbook ──
    global wb
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '日报'
    # 从基准表加载样式模板（无基准时用默认样式）
    header_t = None
    data_t = None
    if last_xlsx:
        header_t, data_t = load_template(last_xlsx)
    write_header(ws, styles, header_t)

    row = 2
    seq = max_seq
    repo_prev = None

    # ── 4. 复制上月未完成任务（B/C/D/E/F/G/H/N 列原样保留，仅 A 列记当天）──
    copy_count = 0
    for u in unfinished:
        seq = u['seq']
        if seq > max_seq:
            max_seq = seq
        repo_name = u['repo'] if u['repo'] != repo_prev else ''
        if repo_name:
            repo_prev = repo_name

        write_row(ws, row, seq, repo_name, u['desc'], u['pct'], u['hours'],
                  u['g_date'] or TODAY, u['ai_note'], styles, f_date=u['f_date'], tdata=data_t)
        print(f"  复制未完成 Row{row}: [{repo_prev}] #{seq} G={u['g_date']} H={u['hours']}h")
        row += 1
        copy_count += 1

    # ── 5. 追加当天 md 新任务（G 列从最后一个未完成任务的 G 列继续叠加）──
    # 基准只算一次，循环内逐任务延续 current_g/current_remaining（与 daily.py 链式一致）
    if unfinished:
        last_u = unfinished[-1]
        gd = last_u['g_date'] or TODAY
        # 计算该 G 日期已有工时
        day_hours = sum(u2['hours'] for u2 in unfinished if u2['g_date'] and u2['g_date'].date() == gd.date())
        gr = day_hours % 8
    else:
        gd = TODAY
        gr = 0

    for t in new_tasks:
        seq += 1
        repo_name = t['repo'] if t['repo'] != repo_prev else ''
        if repo_name:
            repo_prev = repo_name

        gr2 = gr + t['human_h']
        gd2 = gd
        while gr2 > 8:
            gr2 -= 8
            gd2 = next_workday(gd2)

        ai_note = f"预估AI辅助工时(h)：{t['ai_h']}" if t['ai_h'] > 0 else ''
        write_row(ws, row, seq, repo_name, t['desc'], '0%', t['human_h'], gd2, ai_note, styles, tdata=data_t)
        gd = gd2; gr = gr2
        print(f"  新增 Row{row}: [{repo_prev}] #{seq} G={gd2.strftime('%m-%d')} H={t['human_h']}h")
        row += 1

    # ── 6. 备注 + 收尾 ──
    write_notes(ws, copy_count + len(new_tasks), styles)
    finalize(ws, 2, row - 1)
    print(f'Done: {XL}')


if __name__ == '__main__':
    main()
